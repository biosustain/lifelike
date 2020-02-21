import attr

from typing import Dict, List, NamedTuple, Optional, Union

from py2neo import cypher, NodeMatcher, RelationshipMatcher
from werkzeug.datastructures import FileStorage

from neo4japp.services.common import BaseDao
from neo4japp.models import GraphNode, GraphRelationship
from neo4japp.constants import *
from neo4japp.factory import cache
from neo4japp.util import CamelDictMixin, compute_hash
from neo4japp.services.common import BaseDao

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from py2neo import (
    Graph,
    Node,
    Transaction,
    NodeMatch,
    Relationship,
    RelationshipMatch,
    RelationshipMatcher,
)

@attr.s(frozen=True)
class FileNameAndSheets(CamelDictMixin):
    @attr.s(frozen=True)
    class SheetNameAndColumnNames(CamelDictMixin):
        sheet_name: str = attr.ib()
        # key is column name, value is column index
        sheet_column_names: List[Dict[str, int]] = attr.ib()

    sheets: List[SheetNameAndColumnNames] = attr.ib()
    filename: str = attr.ib()


@attr.s(frozen=True)
class Neo4jColumnMapping(CamelDictMixin):
    """The int values are the column index
    from the excel files."""
    @attr.s(frozen=True)
    class Neo4jNodeMapping(CamelDictMixin):
        node_type: str = attr.ib()
        node_properties: Dict[int, str] = attr.ib()
        mapped_node_type: str = attr.ib()
        mapped_node_property: str = attr.ib()
        # this will be used to match a node
        unique_property: str = attr.ib()

    @attr.s(frozen=True)
    class Neo4jRelationshipMapping(CamelDictMixin):
        @attr.s(frozen=True)
        class ExistingGraphDBMapping(CamelDictMixin):
            mapped_node_type: str = attr.ib()
            mapped_node_property: Dict[int, str] = attr.ib()

        edge: str = attr.ib()
        edge_property: Dict[int, str] = attr.ib()
        source_node: ExistingGraphDBMapping = attr.ib()
        target_node: ExistingGraphDBMapping = attr.ib()

    file_name: str = attr.ib()
    sheet_name: str = attr.ib()
    node: Optional[Neo4jNodeMapping] = attr.ib(default=None)
    relationship: Optional[Neo4jRelationshipMapping] = attr.ib(default=None)


class Neo4JService(BaseDao):
    def __init__(self, graph):
        super().__init__(graph)

    def _query_neo4j(self, query: str):
        # TODO: Can possibly use a dispatch method/injection
        # of sorts to use custom labeling methods for
        # different type of nodes/edges being converted.
        # The default does not always set an appropriate label
        # name.
        records = self.graph.run(query).data()
        if not records:
            return None
        node_dict = dict()
        rel_dict = dict()
        for record in records:
            nodes = record['nodes']
            rels = record['relationships']
            for node in nodes:
                graph_node = GraphNode.from_py2neo(
                    node,
                    display_fn=lambda x: x.get(DISPLAY_NAME_MAP[next(iter(node.labels), set())])
                )
                node_dict[graph_node.id] = graph_node
            for rel in rels:
                graph_rel = GraphRelationship.from_py2neo(rel)
                rel_dict[graph_rel.id] = graph_rel
        return dict(nodes=[n.to_dict() for n in node_dict.values()],
                    edges=[r.to_dict() for r in rel_dict.values()])


    def get_organisms(self):
        nodes = list(NodeMatcher(self.graph).match(NODE_SPECIES))
        organism_nodes = [
            GraphNode.from_py2neo(n, display_fn=lambda x: x.get('common_name'))
                for n in nodes
        ]
        return dict(nodes=[n.to_dict() for n in organism_nodes], edges=[])

    def get_some_diseases(self):
        nodes = list(NodeMatcher(self.graph).match(TYPE_DISEASE).limit(10))
        disease_nodes = [
            GraphNode.from_py2neo(n, display_fn=lambda x: x.get(DISPLAY_NAME_MAP[TYPE_DISEASE]))
                for n in nodes
        ]
        return dict(nodes=[n.to_dict() for n in disease_nodes], edges=[])

    def get_biocyc_db(self, org_ids: [str]):
        if org_ids:
            query = f'match(n:Species) where n.biocyc_id in {str(org_ids)} return labels(n) as node_labels'
            records = list(self.graph.run(query))
            db_labels = []
            for record in records:
                labels = record['labels']
                for label in labels:
                    if label not in set(DB_BIOCYC, NODE_SPECIES):
                        db_labels.append(label)
            return db_labels
        return None

    def load_regulatory_graph(self, req):
        db_filter = self.get_biocyc_db(req.org_ids)
        if req.is_gene():
            query = self.get_gene_regulatory_query(req, db_filter)
            return self._query_neo4j(query)
        return None

    def expand_graph(self, node_id: str, limit: int):
        query = self.get_expand_query(node_id, limit)
        return self._query_neo4j(query)

    def get_association_sentences(self, node_id: str, description: str, entry_text: str):
        query = self.get_association_sentences_query(node_id, description, entry_text)
        data = self.graph.run(query).data()
        return [result['references'] for result in data]

    def load_reaction_graph(self, biocyc_id: str):
        query = self.get_reaction_query(biocyc_id)
        return self._query_neo4j(query)

    def get_graph(self, req):
        # TODO: Make this filter non-static
        db_filter = self.get_biocyc_db(req.org_ids)
        query = ''
        if req.is_gene():
            query = self.get_gene_gpr_query(req, db_filter)
        elif req.is_protein():
            query = self.get_protein_gpr_query(req, db_filter)
        elif req.is_compound():
            query = self.get_compound_query(req, db_filter)
        elif req.is_pathway():
            query = self.get_pathway_gpr_query(req, db_filter)
        elif req.is_chemical():
            query = self.get_chemical_query(req, db_filter)
        return self._query_neo4j(query)

    def get_db_labels(self) -> List[str]:
        """Get all labels from database."""
        labels = self.graph.run('call db.labels()').data()
        return [label['label'] for label in labels]

    def get_node_properties(self, node_label) -> Dict[str, List[str]]:
        """Get all properties of a label."""
        props = self.graph.run(f'match (n: {node_label}) unwind keys(n) as key return distinct key').data()
        return { node_label: [prop['key'] for prop in props] }

    def get_gene_gpr_query(self, req, db_filter: [str]):
        """NOTE: the `get_node_label` is not to get from the database,
        but from the request object class SearchResult.
        """
        gene_label = req.node_label()
        enz_label = req.get_node_label(TYPE_ENZREACTION, req.get_db_labels())
        args = {PROP_BIOCYC_ID: req.id, 'gene_label': gene_label, 'enz_label': enz_label}
        return """
            match(g{gene_label}) where g.biocyc_id = '{biocyc_id}' with g
            optional match p1 = (g)-[:ENCODES]-(:Protein)-[:HAS_COMPONENT*0..3]-(:Protein) unwind nodes(p1) as prot
            optional match p2=(prot)-[:CATALYZES]->(:EnzReaction)-[:HAS_REACTION]->(r:Reaction)
            optional match p3 = ({gene_label})-[:ENCODES]->(:Protein)<-[:HAS_COMPONENT*0..3]-(:Protein)
            -[:CATALYZES]->(enz{enz_label})-[:HAS_REACTION]->(r)
            return [g]+ COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) + COALESCE(nodes(p3), []) as nodes,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) + COALESCE(relationships(p3), []) as relationships
        """.format(**args)

    def get_protein_gpr_query(self, req, db_filter):
        enz_label = req.get_node_label(TYPE_ENZREACTION, req.get_db_labels())
        args = {PROP_BIOCYC_ID: req.id, 'protein_label': req.node_label(), 'enz_label': enz_label}
        return """
            match p1 = ({protein_label})-[:HAS_COMPONENT*0..3]->(n:Protein)-[:HAS_COMPONENT*0..3]->(:Protein)
            where n.biocyc_id = '{biocyc_id}'
            unwind nodes(p1) as prot
            optional match p2=(:Gene)-[:ENCODES]->(prot)
            optional match rp=(prot)-[:CATALYZES]->(:EnzReaction)-[:HAS_REACTION]->(r)
            optional match p3 = (:Gene:Ecocyc)-[:ENCODES]->({protein_label})<-[:HAS_COMPONENT*0..3]-
            (:Protein)-[:CATALYZES]->(enz{enz_label})-[:HAS_REACTION]->(r)
            return COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) + COALESCE(nodes(p3), []) as nodes,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) + COALESCE(relationships(p3), []) as relationships
        """.format(**args)

    def get_pathway_gpr_query(self, req, db_filter:[str]):
        args = {PROP_BIOCYC_ID: req.id}
        query = """
            match (n:Pathway) where n.biocyc_id = '{biocyc_id}'
            with n
            optional match p1 = (n)-[:CONTAINS]-(:Reaction)-[:HAS_REACTION]-(enz:EnzReaction)-[:CATALYZES]-(prot:Protein)
            optional match p2 = (:Gene)-[:ENCODES]->(:Protein)<-[:HAS_COMPONENT*0..3]-(prot)
            return [n]+ COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) as nodes,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) as relationships
        """.format(**args)
        return query

    def get_compound_query(self, req, db_filter:[str]):
        args = {PROP_BIOCYC_ID: req.id}
        return """
            MATCH (c:Compound) where c.biocyc_id='{biocyc_id}'
            with c
            OPTIONAL MATCH p1=(c)-[:IS_CONSUMED_BY]->(:Reaction)
            OPTIONAL Match p2=(c)-[:PRODUCES]-(:Reaction)
            with [c]+ COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) as n,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) as r
            unwind n as allnodes
            unwind r as allrels
            return collect(distinct allnodes) as nodes, collect(distinct allrels) as relationships
        """.format(**args)

    def get_chemical_query(self, req, db_filter:[str]):
        args = {PROP_CHEBI_ID: req.id}
        return """
            match(c:CHEBI) where c.chebi_id = '{chebi_id}' with c
            optional match p1=(c)-[:IS_A]-(:CHEBI)
            optional match p2=(c)-[:IS_CONJUGATE_ACID_OF]-(:CHEBI)
            optional match p3=(c)-[:HAS_PART]-(:CHEBI)
            return [c]+ COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) + COALESCE(nodes(p3), []) as nodes,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) + COALESCE(relationships(p3), []) as relationships
        """.format(**args)

    def get_gene_regulatory_query(self, req, db_filter:[str]):
        args = {PROP_BIOCYC_ID: req.id}
        return """
            match (n:Gene)-[:IS]-(:NCBI:Gene)-[:IS]-(g:Gene:RegulonDB) where n.biocyc_id = 'EG11530' with g
            optional match p1= (:Gene)-[:ENCODES]->(:Product)-[:IS]-(:TranscriptionFactor)-[:REGULATES]->(g)
            optional match p2 = (g)-[:ENCODES]->(:Product)-[:IS]-(:TranscriptionFactor)-[:REGULATES]->(:Gene)
            return [g]+ COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) as nodes,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) as relationships
        """.format(**args)

    # TODO: Allow flexible limits on nodes; enable this in the blueprints
    def get_expand_query(self, node_id: str, limit: int = 50):
        query = """
            MATCH (n)-[l:ASSOCIATED]-(s) WHERE ID(n) = {}
            WITH n, s, l
            LIMIT {}
            return collect(n) + collect(s) as nodes, collect(l) as relationships
        """.format(node_id, limit)
        return query

    # TODO: Need to make a solid version of this query, not sure the current
    # iteration will work 100% of the time
    def get_association_sentences_query(self, node_id: str, description: str, entry_text: str):
        query = """
            MATCH (n)-[:HAS_ASSOCIATION]-(s:Association)-[:HAS_REF]-(r:Reference)
            WHERE ID(n) = {} AND s.description='{}' AND r.entry2_text=~'.*{}.*'
            WITH r
            return r as references
        """.format(node_id, description, entry_text)
        return query

    def get_reaction_query(self, biocyc_id: str):
        query = """
            match(n:Reaction) where n.biocyc_id = '{}' with n
            optional match p1 = (n)-[]-(:EnzReaction)-[:CATALYZES]-(prot:Protein)
            optional match p2 = (:Gene)-[:ENCODES]->(:Protein)<-[:HAS_COMPONENT*0..3]-(prot)
            return [n]+ COALESCE(nodes(p1), [])+ COALESCE(nodes(p2), []) as nodes,
            COALESCE(relationships(p1), []) + COALESCE(relationships(p2), []) as relationships
        """.format(biocyc_id)
        print(query)
        return query

    def parse_file(self, f: FileStorage) -> Workbook:
        workbook = load_workbook(f)
        print(f'caching {f.filename} as key')
        cache.set(f.filename, workbook)
        return workbook

    def get_workbook_sheet_names_and_columns(
        self,
        filename: str,
        workbook: Workbook,
    ) -> FileNameAndSheets:
        sheet_list = []
        for i in range(0, len(workbook.sheetnames)):
            workbook.active = i
            current_ws: Worksheet = workbook.active

            sheet_col_names = []
            # loop through just the first row
            for i, cols in enumerate(current_ws.iter_cols(
                max_row=1,
                max_col=len(list(current_ws.columns)),
                values_only=True,
            )):
                if cols[0]:
                    sheet_col_names.append({cols[0]: i})

            sheet_list.append(FileNameAndSheets.SheetNameAndColumnNames(
                sheet_name=current_ws.title,
                sheet_column_names=sheet_col_names,
            ))
        sheets = FileNameAndSheets(sheets=sheet_list, filename=filename)
        return sheets

    def create_node_mapping(
        self,
        column_mappings: Neo4jColumnMapping,
    ) -> dict:
        def create_node(unique_prop, row, props, node_type):
            node = {}  # type: ignore
            node_props = {}

            # value becomes the key representing property label
            # key is used to get the value in excel
            # e.g { prop_label_in_db: cell_value_in_excel_file }
            for k, v in props.items():
                if type(row[k]) is str:
                    node_props[v] = row[k].strip().lstrip()
                else:
                    node_props[v] = row[k]

            node['data'] = node_props
            node['label'] = node_type  # type: ignore
            node['unique_property'] = {unique_prop: node_props[unique_prop]}
            node['hash'] = compute_hash(node)  # type: ignore
            return node

        nodes = []

        col_idx_node_prop = {}

        # TODO: why is the key changing to str?
        for k, v in column_mappings.node.node_properties.items():
            col_idx_node_prop[int(k)] = v

        workbook = cache.get(column_mappings.file_name)

        for i, name in enumerate(workbook.sheetnames):
            if name == column_mappings.sheet_name:
                workbook.active = i
                break

        current_ws = workbook.active

        # unmerge any cells first
        # and assign the value to them
        for group in current_ws.merged_cell_ranges:
            min_col, min_row, max_col, max_row = group.bounds
            value_to_assign = current_ws.cell(row=min_row, column=min_col).value
            current_ws.unmerge_cells(str(group))
            for row in current_ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = value_to_assign

        for row in current_ws.iter_rows(
            min_row=2,
            max_row=len(list(current_ws.rows)),
            max_col=len(list(current_ws.columns)),
            values_only=True,
        ):
            node_type = column_mappings.node.node_type
            if not all(cell is None for cell in row):
                nodes.append(create_node(
                    unique_prop=column_mappings.node.unique_property,
                    row=row,
                    props=col_idx_node_prop,
                    node_type=node_type),
                )
        return {'nodes': nodes}

    def search_for_relationship_with_label_and_properties(
        self,
        label: str,
        props: dict,
        src: Node,
        dest: Node,
    ) -> RelationshipMatch:
        if label is None or not label:
            raise Exception(
                'No label provided for search query! Is the JSON object malformed?\n' +
                'Props:\n' +
                ''.join([f'\t{key}: {value}\n' for key, value in props.items()])
            )
        rm = RelationshipMatcher(self.graph)
        return rm.match([src, dest], label, **props).first()

    def goc_relationship_from_data(
        self,
        node_hash_map: dict,
        rel_data: dict,
        tx: Transaction,
    ) -> Relationship:
        label = rel_data['label']
        src_node = node_hash_map[rel_data['src']]
        dest_node = node_hash_map[rel_data['dest']]
        props = rel_data['data']

        existing_relationship = self.search_for_relationship_with_label_and_properties(
            label, props, src_node, dest_node)

        if existing_relationship:
            return existing_relationship
        else:
            new_relationship = Relationship(src_node, label, dest_node, **props)
            tx.create(new_relationship)
            return new_relationship

    def search_for_node_with_label_and_properties(
        self,
        label: str,
        props: dict,
    ) -> NodeMatch:
        if label is None or not label:
            raise Exception(
                'No label provided for search query! Is the JSON object malformed?\n' +
                'Props:\n' +
                ''.join([f'\t{key}: {value}\n' for key, value in props.items()])
            )
        return self.graph.nodes.match(label, **props).first()

    def goc_node_from_data(
        self,
        node_data: dict,
        tx: Transaction,
    ) -> Node:
        label = node_data['label']
        props = node_data['data']

        existing_node = self.search_for_node_with_label_and_properties(label, props)

        if existing_node:
            return existing_node
        else:
            new_node = Node(label, **props)
            tx.create(new_node)
            return new_node

    def save_node_to_neo4j(self, column_mappings: Neo4jColumnMapping) -> None:
        print(column_mappings)
        node_mapping = self.create_node_mapping(column_mappings)

        tx = self.graph.begin()
        node_hash_map = dict()

        # import IPython; IPython.embed()

        # create the experimental nodes
        # TODO: create index on unique_property
        tx.run(
            'unwind {batch} as row ' \
            'call apoc.merge.node([row.label], row.unique_property, row.data, row.data) ' \
            'yield node ' \
            'return count(*)',
            {'batch': node_mapping['nodes']}
        )

        # import IPython; IPython.embed()

        print('Done creating nodes')

        # create relationship from experimental nodes
        # to universal source nodes
        for row in node_mapping['nodes']:
            filter_label = next(iter(row['unique_property']))
            filter_prop = next(iter(row['unique_property'].values()))
            row_label = row['label']

            query = f'match (universal:`{column_mappings.node.mapped_node_type}` {{`{filter_label}`: "{filter_prop}"}}), '
            query += f'(src:`{row_label}` {{`{filter_label}`: "{filter_prop}"}}) '
            query += f'merge (src)-[:IS_A]->(universal)'
            tx.run(query)

        # import IPython; IPython.embed()

        print('Done creating relationship of new nodes to existing KG')
        tx.commit()

        # # create index on all properties
        # # have to do this here because can't have same transaction
        # # TODO: Jira KG-51
        # for row in node_mapping['nodes']:
        #     label = row['label']
        #     for k, _ in row['data'].items():
        #         tx.run(f'create index on :{label}({k})')
        if column_mappings.relationship.edge:
            self.save_relationship_to_neo4j(column_mappings)
        print('Done')

    def save_relationship_to_neo4j(self, column_mappings: Neo4jColumnMapping) -> None:
        workbook = cache.get(column_mappings.file_name)

        col_idx_src_node_prop = {}
        col_idx_tgt_node_prop = {}
        col_idx_edge_prop = {}

        for k, v in column_mappings.relationship.source_node.mapped_node_property.items():
            col_idx_src_node_prop[int(k)] = v

        for k, v in column_mappings.relationship.target_node.mapped_node_property.items():
            col_idx_tgt_node_prop[int(k)] = v

        for k, v in column_mappings.relationship.edge_property.items():
            col_idx_edge_prop[int(k)] = v

        for i, name in enumerate(workbook.sheetnames):
            if name == column_mappings.sheet_name:
                workbook.active = i
                break

        current_ws = workbook.active

        # unmerge any cells first
        # and assign the value to them
        for group in current_ws.merged_cell_ranges:
            min_col, min_row, max_col, max_row = group.bounds
            value_to_assign = current_ws.cell(row=min_row, column=min_col).value
            current_ws.unmerge_cells(str(group))
            for row in current_ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = value_to_assign

        tx = self.graph.begin()

        for row in current_ws.iter_rows(
            min_row=2,
            max_row=len(list(current_ws.rows)),
            max_col=len(list(current_ws.columns)),
            values_only=True,
        ):
            if not all(cell is None for cell in row):
                # TODO: handle null otherwise strip() and lstrip() fails
                src_label = column_mappings.relationship.source_node.mapped_node_type
                # try:
                src_prop_label = next(iter(col_idx_src_node_prop.values()))
                src_prop_value = row[next(iter(col_idx_src_node_prop))]
                if type(src_prop_value) is str:
                    src_prop_value = f'"{src_prop_value.strip().lstrip()}"'

                tgt_label = column_mappings.relationship.target_node.mapped_node_type
                tgt_prop_label = next(iter(col_idx_tgt_node_prop.values()))
                tgt_prop_value = row[next(iter(col_idx_tgt_node_prop))]
                if type(tgt_prop_value) is str:
                    tgt_prop_value = f'"{tgt_prop_value.strip().lstrip()}"'

                # TODO: instead of using column header
                # use column values for edge
                edge_label = column_mappings.relationship.edge

                query = f'match (s:`{src_label}` {{`{src_prop_label}`: {src_prop_value}}}), '
                query += f'(t:`{tgt_label}` {{`{tgt_prop_label}`: {tgt_prop_value}}}) '

                query += f'merge (s)-[:`{edge_label}`'
                if col_idx_edge_prop:
                    query += f'{{'
                    for k, v in col_idx_edge_prop.items():
                        prop_value = row[k]
                        if type(prop_value) is int:
                            query += f'`{v}`: {prop_value}, '
                        else:
                            query += f'`{v}`: "{prop_value}", '
                    # remove the comma at the end
                    query = query[:-2] + f'}}'
                query += f']->(t)'
                print(query)
                tx.run(query)
                # except AttributeError:
                    # if this occur then the cell in the row
                    # was empty so strip() and lstrip() failed
                    # skip and don't create the relationship
                    # continue
        tx.commit()
        print('Done creating relationships between new nodes')
