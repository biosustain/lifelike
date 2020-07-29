from math import floor

from typing import (
    Any,
    Dict,
    List,
    Tuple
)

from werkzeug.datastructures import FileStorage

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from py2neo import (
    Node,
    Transaction,
    Relationship,
    cypher_escape,
    cypher_repr
)

from neo4japp.factory import cache
from neo4japp.data_transfer_objects.user_file_import import (
    FileNameAndSheets,
    GraphCreationMapping,
    GeneImportRelationship,
    GraphNodeCreationMapping,
    GeneMatchingProperty,
    GraphRelationshipCreationMapping,
    Neo4jColumnMapping,
    Properties,
)
from neo4japp.services.common import GraphBaseDao
from neo4japp.util import compute_hash


class UserFileImportService(GraphBaseDao):
    def __init__(self, graph):
        super().__init__(graph)

    # def strip_leading_and_trailing_spaces(self, current_ws: Worksheet) -> Worksheet:
    #     # remove leading and trailing space
    #     curr_row = 1
    #     max_row = len(list(current_ws.rows))
    #     while curr_row <= max_row:
    #         curr_col = 1
    #         max_col = len(list(current_ws.columns))

    #         while curr_col <= max_col:
    #             value = current_ws.cell(row=curr_row, column=curr_col).value
    #             if type(value) is str:
    #                 current_ws.cell(row=curr_row, column=curr_col).value = value.strip().lstrip()
    #             curr_col += 1
    #         curr_row += 1
    #     return current_ws

    def unmerge_cells(self, current_ws: Worksheet) -> Worksheet:
        # unmerge any cells first
        # and assign the value to them
        for group in current_ws.merged_cell_ranges:
            min_col, min_row, max_col, max_row = group.bounds
            value_to_assign = current_ws.cell(row=min_row, column=min_col).value

            if type(value_to_assign) is str:
                value_to_assign = value_to_assign.strip().lstrip()
            current_ws.unmerge_cells(str(group))
            for row in current_ws.iter_rows(
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            ):
                for cell in row:
                    cell.value = value_to_assign
        return current_ws

    def parse_file(self, f: FileStorage) -> Workbook:
        workbook = load_workbook(f)

        for i in range(0, len(workbook.sheetnames)):
            workbook.active = i
            self.unmerge_cells(workbook.active)
            # slow doing it eagerly
            # self.strip_leading_and_trailing_spaces(workbook.active)
        print(f'caching {f.filename} as key')
        cache.set(f.filename, workbook)
        return workbook

    def get_workbook_sheet_names_and_columns(
        self,
        filename: str,
        workbook: Workbook,
    ) -> FileNameAndSheets:
        sheet_list = []

        for i in range(0, len(workbook.sheetnames)):
            workbook.active = i
            current_ws: Worksheet = workbook.active

            sheet_col_names = []
            col_name_map = {}
            sheet_preview = []

            # loop through just the first row
            for i, cols in enumerate(current_ws.iter_cols(
                max_row=1,
                max_col=len(list(current_ws.columns)),
                values_only=True,
            )):
                if cols[0]:
                    sheet_col_names.append({cols[0]: i})
                    col_name_map[i] = cols[0]

            # generate a small preview of the file
            counter_for_ellipse = 2  # min_row

            for row in current_ws.iter_rows(
                min_row=2,
                max_row=5,
                max_col=len(list(current_ws.columns)),
                values_only=True,
            ):
                if not all(cell is None for cell in row):
                    col_row_mapping = {}
                    for i, cell in enumerate(row):
                        if cell:
                            if counter_for_ellipse == 5:  # max_row
                                col_row_mapping[col_name_map[i]] = '...'
                            else:
                                col_row_mapping[col_name_map[i]] = cell
                    counter_for_ellipse += 1

                    sheet_preview.append(col_row_mapping)

            sheet_list.append(FileNameAndSheets.SheetNameAndColumnNames(
                sheet_name=current_ws.title,
                sheet_column_names=sheet_col_names,
                sheet_preview=sheet_preview,
            ))
        sheets = FileNameAndSheets(sheets=sheet_list, filename=filename)
        return sheets

    def new_rows_for_delimiters(
        self,
        current_ws: Worksheet,
        delimiters: Dict[int, str],
    ) -> Worksheet:
        for mapped_column_idx, delimiter in delimiters.items():
            column_idx = int(mapped_column_idx)+1  # openpyxl is 1-index based, so add 1 to column index from excel  # noqa
            curr_row = 2  # openpyxl is 1-index based, skip header row
            max_row = len(list(current_ws.rows))

            while curr_row <= max_row:
                cell_value = current_ws.cell(row=curr_row, column=column_idx).value
                if cell_value and delimiter in cell_value:
                    cell_values_splitted = cell_value.split(delimiter)
                    # create a new row after current one
                    # for each of the splitted up values
                    new_row = curr_row+1
                    for cell_value_split in cell_values_splitted:
                        current_ws.insert_rows(new_row)

                        # loop through each column and add the values
                        # to new row using current row value
                        for i, cell in enumerate(current_ws[new_row]):
                            curr_working_idx = i+1
                            if curr_working_idx == column_idx:
                                cell.value = cell_value_split
                            else:
                                cell.value = current_ws.cell(row=curr_row, column=curr_working_idx).value  # noqa
                        new_row += 1
                        # increment max_row for each new row added
                        max_row += 1
                    # delete current row since it's values are now copied
                    # to the new rows
                    current_ws.delete_rows(curr_row)
                    max_row -= 1
                    # new_row is at next row to work on now
                    # because it was incremented above
                    # after last newly created row
                    # but minus one because deleted current old row
                    # after copying to new rows
                    curr_row = new_row - 1
                else:
                    curr_row += 1
        print('Done creating new rows for delimiters')
        return current_ws

    def create_node_mappings(
        self,
        column_mappings: Neo4jColumnMapping,
        current_ws: Worksheet,
    ) -> List[GraphNodeCreationMapping]:
        nodes = []
        max_row = len(list(current_ws.rows))

        for node in column_mappings.new_nodes:
            curr_row = 2    # openpyxl is 1-indexed based, so don't count header row
            while curr_row <= max_row:
                # only create nodes that have content
                # skip empty rows in user data file
                if node.node_properties:
                    node_properties = {}
                    for k, v in node.node_properties.items():
                        value = current_ws.cell(row=curr_row, column=int(k)+1).value
                        if type(value) is str:
                            if value.startswith(' ') or value.endswith(' '):
                                value = value.strip().lstrip()
                        if value:
                            node_properties[v] = value

                    cell_value = current_ws.cell(
                        row=curr_row,
                        column=int(next(iter(node.mapped_node_property_from)))+1).value

                    if type(cell_value) is str:
                        if cell_value.startswith(' ') or cell_value.endswith(' '):
                            cell_value = cell_value.strip().lstrip()

                    mapped_node_prop_value = cell_value
                    mapped_node_prop = next(iter(node.mapped_node_property_from.values()))

                    # TODO: Should the attributes of GraphNodeCreationMapping be made optional...?
                    nodes.append(GraphNodeCreationMapping(
                        domain=node.domain,  # type: ignore
                        node_type=node.node_type,  # type: ignore
                        node_properties=node_properties,
                        mapped_node_prop=mapped_node_prop,  # type: ignore
                        mapped_node_prop_value=mapped_node_prop_value,
                        kg_mapped_node_type=node.mapped_node_type,
                        kg_mapped_node_prop=node.mapped_node_property_to,  # type: ignore
                        edge_label=node.edge,  # type: ignore
                    ))
                curr_row += 1
        return nodes

    def create_relationship_mappings(
        self,
        column_mappings: Neo4jColumnMapping,
        current_ws: Worksheet,
    ) -> List[GraphRelationshipCreationMapping]:
        relationships = []
        max_row = len(list(current_ws.rows))

        for relation in column_mappings.relationships:
            curr_row = 2    # openpyxl is 1-indexed based, so don't count header row
            while curr_row <= max_row:
                # if edge int key is negative that means user created a new edge
                # TODO: edge property
                edge_col_idx = int(next(iter(relation.edge)))
                if edge_col_idx == -1:
                    edge_label = next(iter(relation.edge.values()))
                else:
                    edge_label = current_ws.cell(
                        row=curr_row,
                        column=edge_col_idx+1).value

                # need to use node_properties as filter for node
                # because unique_property might not be unique
                # TODO: how to handle this unique_property not being unique?
                # user chooses the column...
                # same issue with target_node - the mapped column might not be unique
                source_node_properties = {}
                for k, v in relation.source_node.node_properties.items():
                    value = current_ws.cell(row=curr_row, column=int(k)+1).value
                    if type(value) is str:
                        if value.startswith(' ') or value.endswith(' '):
                            value = value.strip().lstrip()
                    source_node_properties[v] = value

                # for source node, we know unique property
                # and mapped_node_property_from, so use that as filter to find node
                source_node_label = relation.source_node.node_type
                source_node_prop_label = relation.source_node.unique_property
                cell_value = current_ws.cell(
                    row=curr_row,
                    column=int(next(iter(relation.source_node.mapped_node_property_from)))+1).value

                if type(cell_value) is str:
                    if cell_value.startswith(' ') or cell_value.endswith(' '):
                        cell_value = cell_value.strip().lstrip()
                source_node_prop_value = cell_value

                # for target node, mapped_node_property_to is the
                # target node label, and the int key for
                # mapped_node_property_from is the column to get value
                if relation.target_node.mapped_to_universal_graph:
                    # get the node label from mapped_to_universal_graph dict instead
                    target_node_label = relation.target_node.mapped_to_universal_graph.universal_graph_node_type  # noqa
                    target_node_prop_label = relation.target_node.mapped_to_universal_graph.universal_graph_node_property_label  # noqa
                else:
                    target_node_label = relation.target_node.mapped_node_type
                    target_node_prop_label = relation.target_node.mapped_node_property_to   # type: ignore  # noqa
                target_node_prop_value = current_ws.cell(
                    row=curr_row, column=int(next(iter(relation.target_node.mapped_node_property_from)))+1).value  # noqa

                relationships.append(GraphRelationshipCreationMapping(
                    source_node_label=source_node_label,  # type: ignore
                    source_node_prop_label=source_node_prop_label,  # type: ignore
                    source_node_prop_value=source_node_prop_value,
                    source_node_properties=source_node_properties,
                    target_node_label=target_node_label,
                    target_node_prop_label=target_node_prop_label,
                    target_node_prop_value=target_node_prop_value,
                    edge_label=edge_label,
                ))
                curr_row += 1
        return relationships

    def create_graph_db_mappings(
        self,
        column_mappings: Neo4jColumnMapping,
    ) -> GraphCreationMapping:
        workbook = cache.get(column_mappings.file_name)

        for i, name in enumerate(workbook.sheetnames):
            if name == column_mappings.sheet_name:
                workbook.active = i
                break

        current_ws = workbook.active

        if column_mappings.delimiters:
            # delimiters are for current worksheet via column_mappings.sheet_name
            current_ws = self.new_rows_for_delimiters(current_ws, column_mappings.delimiters)

        nodes_to_create = self.create_node_mappings(column_mappings, current_ws)
        relationships_to_create = self.create_relationship_mappings(column_mappings, current_ws)

        return GraphCreationMapping(
            new_nodes=nodes_to_create, new_relationships=relationships_to_create)

    def save_node_to_neo4j(self, node_mappings: GraphCreationMapping) -> None:
        tx = self.graph.begin()

        # needed because haven't committed yet
        # so the match would not return a node
        created_nodes = set()  # type: ignore
        created_domains = dict()  # type: ignore

        for node in node_mappings.new_nodes:
            # can't use cipher parameters due to the dynamic map keys in filtering
            # e.g merge (n:TYPE {map...})
            # neo4j guy: https://stackoverflow.com/a/28784921
            node_type = node.node_type
            mapped_node_prop = node.mapped_node_prop
            mapped_node_prop_value = node.mapped_node_prop_value
            kg_mapped_node_prop = node.kg_mapped_node_prop
            node_properties = node.node_properties
            kg_mapped_node_type = node.kg_mapped_node_type
            edge_label = node.edge_label

            filter_property = {mapped_node_prop: mapped_node_prop_value}
            kg_filter_property = {kg_mapped_node_prop: mapped_node_prop_value}

            domain_name = node.domain

            if domain_name not in created_domains:
                domain_node = self.graph.nodes.match(domain_name, **{'name': domain_name}).first()

                if not domain_node:
                    domain_node = Node(domain_name, **{'name': domain_name})
                    tx.create(domain_node)
                created_domains[domain_name] = domain_node
            else:
                domain_node = created_domains[domain_name]

            if mapped_node_prop_value not in created_nodes:
                created_nodes.add(mapped_node_prop_value)
                # user experimental data node
                # TODO: currently filter_property is all node properties
                # of a node because there are no unique constraints
                # and we can't tell which "column" a user expects to
                # be unique versus what we have as unique in our database
                exp_node = self.graph.nodes.match(node_type, **filter_property).first()

                if exp_node:
                    # TODO: check if node properties match incoming node_properties
                    # also do in frontend - keep set of values from unique column
                    # if see again, check if node_properties are different
                    # if not, throw exception to let user know (check JIRA issue LL-81)
                    # but what if user wants to create duplicate - would they?
                    continue
                else:
                    exp_node = Node(node_type, **node_properties)
                    tx.create(exp_node)

                # create relationship between user experimental data node with
                # domain node
                relationship = Relationship(domain_node, 'CONTAINS', exp_node, **{})
                tx.create(relationship)

                # create relationship between user experimental data node with
                # existing nodes in knowledge graph
                if kg_mapped_node_type:
                    kg_node = self.graph.nodes.match(kg_mapped_node_type, **kg_filter_property).first()  # noqa
                    if kg_node:
                        relationship = Relationship(exp_node, edge_label, kg_node, **{})
                        tx.create(relationship)

        tx.commit()
        print('Done creating relationship of new nodes to existing KG')
        print('Done creating nodes')

    def save_relationship_to_neo4j(self, node_mappings: GraphCreationMapping) -> None:
        tx = self.graph.begin()

        for relation in node_mappings.new_relationships:
            source_node_label = relation.source_node_label
            source_node_prop_label = relation.source_node_prop_label
            source_node_prop_value = relation.source_node_prop_value
            target_node_label = relation.target_node_label
            target_node_prop_label = relation.target_node_prop_label
            target_node_prop_value = relation.target_node_prop_value
            edge_label = relation.edge_label

            # source_filter_property = {source_node_prop_label: source_node_prop_value}
            # TODO: need to use node_properties here because the filter
            # might not be unique also see above in save_node_to_neo4j()
            source_filter_property = relation.source_node_properties
            target_filter_property = {target_node_prop_label: target_node_prop_value}

            # TODO: if nodes not found throw exception or create?
            # TODO: LL-81: if user selects a column to be a node property
            # it is possible the node label appears multiple times in different rows
            # and the column that is the node property have different values
            # in that case, won't get a source_node here
            # probably fix is to create new node
            # referenced above already (search for LL-81)
            source_node = self.graph.nodes.match(source_node_label, **source_filter_property).first()  # noqa
            if source_node:
                target_node = self.graph.nodes.match(target_node_label, **target_filter_property).first()  # noqa
                if target_node:
                    # TODO: the **{} should be edge properties
                    tx.create(Relationship(source_node, edge_label, target_node, **{}))
        tx.commit()
        print('Done creating relationships between new nodes')

    def get_props_str_from_propnames_and_varname(self, varname: str, propnames: List[str]):
        if (len(propnames) == 0):
            return '{}'

        retval = '{'
        for propname in propnames:
            retval += f'{propname}: {varname}.{propname}, '

        return retval[0:-2] + '}'  # remove trailing ' ,' and add closing curly bracket

    def get_props_obj_from_worksheet(
        self,
        worksheet: Worksheet,
        props: List[Properties],
        row: int
    ):
        props_obj = {}

        for prop in props:
            props_obj[prop.property_name] = worksheet.cell(
                row=row,
                column=int(prop.column) + 1
            ).value or ''

        return props_obj

    def get_merge_worksheet_node_query(self):
        return """
            MERGE (w:Worksheet {name: $worksheet_node_name})
            RETURN ID(w) as worksheet_node_id
        """

    def get_merge_col_match_query(
        self,
        node_label1,
        node_label2,
        node_propnames1,
        node_propnames2,
        rel_label,
        rel_propnames,
    ):
        node_props1 = self.get_props_str_from_propnames_and_varname('tuple[0]', node_propnames1)
        node_props2 = self.get_props_str_from_propnames_and_varname('tuple[1]', node_propnames2)
        rel_props = self.get_props_str_from_propnames_and_varname('tuple[2]', rel_propnames)

        return f"""
        MATCH (w:Worksheet)
        WHERE ID(w)=$worksheet_node_id
        WITH w
        UNWIND $col_match_prop_tuples AS tuple
        MERGE (n:{node_label1}:UserData {node_props1})
        MERGE (m:{node_label2}:UserData {node_props2})
        MERGE (n)-[r:{rel_label} {rel_props}]->(m)
        MERGE (m)-[:IMPORTED_FROM]->(w)
        MERGE (n)-[:IMPORTED_FROM]->(w)
        """

    def get_merge_gene_match_query(
        self,
        node_label1,
        node_propnames1,
        rel_label,
        rel_propnames,
        gene_matching_property,
    ):
        node_props1 = self.get_props_str_from_propnames_and_varname('tuple[0]', node_propnames1)
        rel_props = self.get_props_str_from_propnames_and_varname('tuple[1]', rel_propnames)

        if gene_matching_property == GeneMatchingProperty.NAME.value:
            return f"""
            MATCH (w:Worksheet)
            WHERE ID(w)=$worksheet_node_id
            WITH w
            UNWIND $gene_match_prop_tuples AS tuple
            MERGE (n:{node_label1}:UserData {node_props1})
            MERGE (n)-[:IMPORTED_FROM]->(w)
            WITH n
            MATCH (g:Gene)-[:HAS_TAXONOMY]->(t:Taxonomy)
            WHERE g.name=n.cell_value AND ID(t)=$tax_id
            MERGE (n)-[r:{rel_label} {rel_props}]->(g)
            """
        else:
            return f"""
            MATCH (w:Worksheet)
            WHERE ID(w)=$worksheet_node_id
            WITH w
            UNWIND $gene_match_prop_tuples AS tuple
            MERGE (n:{node_label1}:UserData {node_props1})
            MERGE (n)-[:IMPORTED_FROM]->(w)
            WITH n
            MATCH (g:Gene)-[:HAS_TAXONOMY]->(t:Taxonomy)
            WHERE g.id=n.cell_value AND ID(t)=$tax_id
            MERGE (n)-[r:{rel_label} {rel_props}]->(g)
            """

    def get_import_batches(
        self,
        data,
    ):
        num_batches = floor((len(data) / 512)) + 1
        batches = []
        for i in range(0, num_batches):
            batch = data[i * 512:(i + 1) * 512]
            batches.append(batch)
        return batches

    def get_worksheet(self, file_name: str, sheet_name: str):
        workbook = cache.get(file_name)

        for i, name in enumerate(workbook.sheetnames):
            if name == sheet_name:
                workbook.active = i
                break

        return workbook.active

    def import_gene_relationships(
        self,
        file_name: str,
        sheet_name: str,
        worksheet_node_name: str,
        relationships: List[GeneImportRelationship],
    ):
        worksheet = self.get_worksheet(file_name, sheet_name)
        max_row = len(list(worksheet.rows))
        curr_row = 2  # start at the second row because we don't want to include headers

        relationship_hashes = []
        rel_hash_map = {}  # {relhash: relationship}
        col_match_prop_tuples = {}  # {rel_hash: [(node1, node2, relationship_props)]}
        gene_match_prop_tuples = {}  # {rel_hash: [(node1, relationship_props)]}

        # Setup hash maps
        for relationship in relationships:
            rel_hash = compute_hash(relationship.to_dict())
            relationship_hashes.append(rel_hash)

            rel_hash_map[rel_hash] = relationship

            if col_match_prop_tuples.get(rel_hash, None) is None:
                col_match_prop_tuples[rel_hash] = []
            if gene_match_prop_tuples.get(rel_hash, None) is None:
                gene_match_prop_tuples[rel_hash] = []

        # Setup lists of property tuples by reading data from the spreadsheet.
        while curr_row <= max_row:
            for rel_hash in relationship_hashes:
                relationship = rel_hash_map[rel_hash]
                relationship_props = self.get_props_obj_from_worksheet(
                    worksheet=worksheet,
                    props=relationship.relationship_properties,
                    row=curr_row,
                )

                # openpyxl is 1-based, so add 1 to the column_index
                node1_val = worksheet.cell(row=curr_row, column=int(relationship.column_index1)+1).value  # noqa
                node_props1 = {
                    'cell_value': node1_val or '',
                }
                node_props1.update(
                    self.get_props_obj_from_worksheet(
                        worksheet=worksheet,
                        props=relationship.node_properties1,
                        row=curr_row,
                    )
                )

                # If either species_selection or gene_matching_property are null, then this
                # relationship is a column-to-column relationship.
                if relationship.species_selection is None or relationship.gene_matching_property is None:  # noqa
                    node2_val = worksheet.cell(row=curr_row, column=int(relationship.column_index2)+1).value  # noqa
                    node_props2 = {
                        'cell_value': node2_val or '',
                    }
                    node_props2.update(
                        self.get_props_obj_from_worksheet(
                            worksheet=worksheet,
                            props=relationship.node_properties2,
                            row=curr_row,
                        )
                    )
                    col_match_prop_tuples[rel_hash].append(
                        (node_props1, node_props2, relationship_props)
                    )
                else:
                    gene_match_prop_tuples[rel_hash].append(
                        (node_props1, relationship_props)
                    )
            curr_row += 1

        tx = self.graph.begin()

        # Merge (get or create) the worksheet node, and get the ID of the merged node
        merge_worksheet_query = self.get_merge_worksheet_node_query()
        worksheet_node_id = tx.run(
            merge_worksheet_query,
            {
                'worksheet_node_name': worksheet_node_name
            }
        ).evaluate()

        for rel_hash in relationship_hashes:
            relationship = rel_hash_map[rel_hash]
            relationship_label = relationship.relationship_label
            relationship_propnames = [
                prop.property_name for prop in relationship.relationship_properties
            ]
            node_label1 = relationship.node_label1
            node_label2 = relationship.node_label2

            # If this relationship describes a column-to-column mapping, merge the two new nodes
            # and map them to each other.
            if relationship.species_selection is None or relationship.gene_matching_property is None:  # noqa
                merge_col_match_query = self.get_merge_col_match_query(
                    node_label1,
                    node_label2,
                    [prop.property_name for prop in relationship.node_properties1] + ['cell_value'],  # noqa
                    [prop.property_name for prop in relationship.node_properties2] + ['cell_value'],  # noqa
                    relationship_label,
                    relationship_propnames,
                )

                # Running the merges in batches gives a noticeable performance increase
                batches = self.get_import_batches(col_match_prop_tuples[rel_hash])
                for batch in batches:
                    tx.run(
                        merge_col_match_query,
                        {
                            'col_match_prop_tuples': batch,
                            'worksheet_node_id': worksheet_node_id
                        }
                    )
            # If this relationship describes a column-to-KG-gene mapping, merge the column node,
            # and map it to the corresponding KG node.
            else:
                merge_gene_match_query = self.get_merge_gene_match_query(
                    node_label1,
                    [prop.property_name for prop in relationship.node_properties1] + ['cell_value'],  # noqa
                    relationship_label,
                    relationship_propnames,
                    relationship.gene_matching_property
                )

                # Running the merges in batches gives a noticeable performance increase
                batches = self.get_import_batches(gene_match_prop_tuples[rel_hash])
                for batch in batches:
                    tx.run(
                        merge_gene_match_query,
                        {
                            'gene_match_prop_tuples': batch,
                            'tax_id': int(relationship.species_selection),
                            'worksheet_node_id': worksheet_node_id
                        }
                    )
        tx.commit()

        return worksheet_node_id
