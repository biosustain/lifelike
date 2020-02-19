import attr

from typing import List, Dict

from werkzeug.datastructures import FileStorage

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from neo4japp.factory import cache
from neo4japp.services.neo4j_service import Neo4jColumnMapping
from neo4japp.util import CamelDictMixin


@attr.s(frozen=True)
class FileNameAndSheets(CamelDictMixin):
    @attr.s(frozen=True)
    class SheetNameAndColumnNames(CamelDictMixin):
        sheet_name: str = attr.ib()
        # key is column name, value is column index
        sheet_column_names: List[Dict[str, int]] = attr.ib()
        sheet_preview: List[Dict[str, str]] = attr.ib()

    sheets: List[SheetNameAndColumnNames] = attr.ib()
    filename: str = attr.ib()

class ImporterService():
    def parse_file(self, f: FileStorage) -> Workbook:
        workbook = load_workbook(f)
        print(f'caching {f.filename} as key')
        cache.set(f.filename, workbook)
        return workbook

    def unmerge_cells(self, current_ws):
        # unmerge any cells first
        # and assign the value to them
        for group in current_ws.merged_cell_ranges:
            min_col, min_row, max_col, max_row = group.bounds
            value_to_assign = current_ws.cell(row=min_row, column=min_col).value
            current_ws.unmerge_cells(str(group))
            for row in current_ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = value_to_assign
        return current_ws

    def get_workbook_sheet_names_and_columns(
        self,
        filename: str,
        workbook: Workbook,
    ) -> FileNameAndSheets:
        sheet_list = []

        for i in range(0, len(workbook.sheetnames)):
            workbook.active = i
            current_ws: Worksheet = self.unmerge_cells(workbook.active)

            sheet_col_names = []
            col_name_map = {}
            sheet_preview = []

            # loop through just the first row
            for i, cols in enumerate(current_ws.iter_cols(
                max_row=1,
                max_col=len(list(current_ws.columns)),
                values_only=True,
            )):
                if cols[0]:
                    sheet_col_names.append({cols[0]: i})
                    col_name_map[i] = cols[0]

            # generate a small preview of the file
            counter_for_ellipse = 2  # min_row

            for row in current_ws.iter_rows(
                min_row=2,
                max_row=8,
                max_col=len(list(current_ws.columns)),
                values_only=True,
            ):
                if not all(cell is None for cell in row):
                    col_row_mapping = {}
                    for i, cell in enumerate(row):
                        if cell:
                            if counter_for_ellipse == 8:  # max_row
                                col_row_mapping[col_name_map[i]] = '...'
                            else:
                                col_row_mapping[col_name_map[i]] = cell
                    counter_for_ellipse += 1

                    sheet_preview.append(col_row_mapping)

            sheet_list.append(FileNameAndSheets.SheetNameAndColumnNames(
                sheet_name=current_ws.title,
                sheet_column_names=sheet_col_names,
                sheet_preview=sheet_preview,
            ))
        sheets = FileNameAndSheets(sheets=sheet_list, filename=filename)
        return sheets

    def create_node_mapping(
        self,
        column_mappings: Neo4jColumnMapping,
    ) -> list:
        workbook = cache.get(column_mappings.file_name)

        for i, name in enumerate(workbook.sheetnames):
            if name == column_mappings.sheet_name:
                workbook.active = i
                break

        current_ws = workbook.active

        nodes = []
        max_row = len(list(current_ws.rows))
        # import IPython; IPython.embed()
        for node in column_mappings.new_nodes:
            curr_row = 2    # openpyxl is 1-indexed based, so don't count header row
            while curr_row <= max_row:
                node_properties = {}
                for k, v in node.node_properties.items():
                    value = current_ws.cell(row=curr_row, column=int(k)+1).value
                    if type(value) is str:
                        value = value.strip().lstrip()
                    node_properties[v] = value

                unstripped_value = current_ws.cell(
                    row=curr_row,
                    column=int(next(iter(node.mapped_node_property_from)))+1).value
                if type(unstripped_value) is str:
                    stripped_value = unstripped_value.strip().lstrip()
                mapped_node_prop_value = stripped_value
                mapped_node_prop = next(iter(node.mapped_node_property_from.values()))

                # TODO: create a attr class for this
                nodes.append({
                    'domain': column_mappings.domain,
                    'node_type': node.node_type,
                    'node_properties': node_properties,
                    'mapped_node_prop': mapped_node_prop,
                    'mapped_node_prop_value': mapped_node_prop_value,
                    'KG_mapped_node_type': node.mapped_node_type,
                    'KG_mapped_node_prop': node.mapped_node_property_to,
                    'edge': node.edge,
                })
                curr_row += 1
        # import IPython; IPython.embed()

        relationships = []
        for relation in column_mappings.relationships:
            curr_row = 2    # openpyxl is 1-indexed based, so don't count header row
            while curr_row <= max_row:
                # TODO: if edge int key is negative that means user created a new edge
                # TODO: edge property
                edge_label = current_ws.cell(
                    row=curr_row,
                    column=int(next(iter(relation.edge)))+1).value

                # need to use node_properties as filter
                # because unique_property might not be unique
                # TODO: how to handle this unique_property not being unique?
                # user chooses the column...
                # same issue with target_node - the mapped column might not be unique
                source_node_properties = {}
                for k, v in relation.source_node.node_properties.items():
                    value = current_ws.cell(row=curr_row, column=int(k)+1).value
                    if type(value) is str:
                        value = value.strip().lstrip()
                    source_node_properties[v] = value

                # for source node, we know unique property
                # and mapped_node_property_from, so use that as filter to find node
                source_node_label = relation.source_node.node_type
                source_node_prop_label = relation.source_node.unique_property
                source_node_prop_value = current_ws.cell(
                    row=curr_row,
                    column=int(next(iter(relation.source_node.mapped_node_property_from)))+1).value

                # for target node, mapped_node_property_to is the
                # target node label, and the int key for
                # mapped_node_property_from is the column to get value
                target_node_label = relation.target_node.mapped_node_type
                target_node_prop_label = relation.target_node.mapped_node_property_to
                target_node_prop_value = current_ws.cell(
                    row=curr_row,
                    column=int(next(iter(relation.target_node.mapped_node_property_from)))+1).value

                # TODO: make attrs class for this
                relationships.append({
                    'source_node_label': source_node_label,
                    'source_node_prop_label': source_node_prop_label,
                    'source_node_prop_value': source_node_prop_value,
                    'source_node_properties': source_node_properties,
                    'target_node_label': target_node_label,
                    'target_node_prop_label': target_node_prop_label,
                    'target_node_prop_value': target_node_prop_value,
                    'edge_label': edge_label,
                })
                curr_row += 1
            # print(relationships)
            # import IPython; IPython.embed()
        return {'new_nodes': nodes, 'relationships': relationships}
